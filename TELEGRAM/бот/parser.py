import sys
import schedule
import os
from telebot import types
from openpyxl import load_workbook
import string

def pars():

    # ПРОВЕРКА НА ВХОИЖДЕНИЕ НОВОГО ФАЙЛА В ДИРЕКТОРИЮ
    path = 'C:/Users/aleks/Desktop/TELEGRAM/РАБОЧЕЕ РАСПИСАНИЕ  2022-2023/Новая папка' # Путь к вашей папке

    # Получим список имен всего содержимого папки
    # и превратим их в абсолютные пути
    dir_list = [os.path.join(path, x) for x in os.listdir(path)]

    if dir_list:
        # Создадим список из путей к файлам и дат их создания.
        date_list = [[x, os.path.getctime(x)] for x in dir_list]

        # Отсортируем список по дате создания в обратном порядке
        sort_date_list = sorted(date_list, key=lambda x: x[0], reverse=True)

        # Выведем первый элемент списка. Он и будет самым последним по дате
        a = (sort_date_list[0][0]).replace("\\", "/")
    # КОНЕЦ ПРОВЕРКИ ФАЙЛА


        path_1 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_101Б.txt'
        path_2 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_101А.txt'
        path_3 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_100.txt'
        path_4 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_200.txt'
        path_5 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_300.txt'
        path_6 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_103.txt'
        path_7 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_203.txt'
        path_8 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_303.txt'
        path_9 = 'C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_102.txt'
        path_10 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_202.txt'
        path_11 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_302.txt'
        path_12 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_201А.txt'
        path_13 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_201Б.txt'
        path_14 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_301А.txt'
        path_15 ='C:/Users/aleks/Desktop/TELEGRAM/Новая папка/www_301Б.txt'


        sys.stdout = open(path_1,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_1 курс']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        sys.stdout = open(path_2,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_1 курс']

        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        sys.stdout = open(path_3,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['физра']

        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        sys.stdout = open(path_4,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['физра']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        sys.stdout = open(path_5,'w')
        book = load_workbook(filename = a)

        sheet_1 = book['физра']

        print(str(sheet_1['K' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))
        for i in range(40,50):
           print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        sys.stdout = open(path_6,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['информатики']

        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        sys.stdout = open(path_7,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['информатики']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))
        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        sys.stdout = open(path_8,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['информатики']

        print(str(sheet_1['K' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))
        for i in range(40,50):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        sys.stdout = open(path_9,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['дошкольное']

        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))



        sys.stdout = open(path_10,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['дошкольное']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        sys.stdout = open(path_11,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['дошкольное']

        print(str(sheet_1['K' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))
        for i in range(40,50):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['J' + str(i)].value).replace("None",""),str(sheet_1['K' + str(i)].value).replace("None",""), str(sheet_1['L' + str(i)].value).replace("None",""), str(sheet_1['M' + str(i)].value).replace("None",""))


        sys.stdout = open(path_12,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_2 курс']

        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        sys.stdout = open(path_13,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_2 курс']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        sys.stdout = open(path_14,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_3 курс']
        print(str(sheet_1['C' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))
        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['B' + str(i)].value).replace("None",""),str(sheet_1['C' + str(i)].value).replace("None",""), str(sheet_1['D' + str(i)].value).replace("None",""), str(sheet_1['E' + str(i)].value).replace("None",""))


        sys.stdout = open(path_15,'w')

        book = load_workbook(filename = a)

        sheet_1 = book['начальные_3 курс']

        print(str(sheet_1['G' + str(9)].value))
        print("понедельник")
        for i in range(10,20):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("вторник")
        for i in range(20,30):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))


        print("среда")
        for i in range(30,40):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("четверг")
        for i in range(40,50):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("пятница")
        for i in range(50,60):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))

        print("суббота")
        for i in range(60,66):
            print(str(sheet_1['F' + str(i)].value).replace("None",""),str(sheet_1['G' + str(i)].value).replace("None",""), str(sheet_1['H' + str(i)].value).replace("None",""), str(sheet_1['I' + str(i)].value).replace("None",""))



def main():
    schedule.every().saturday.at('14:00').do(pars)
    while True:
        schedule.run_pending()



if __name__ == '__main__':
    main()